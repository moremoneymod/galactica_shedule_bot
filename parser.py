import os.path

from xls2xlsx import XLS2XLSX
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict
import re


class SheduleParser:
    def __init__(self, document_path: str) -> None:
        self._document_path = document_path

    def _open_worksheet(self) -> None:
        if self._document_path.endswith(".xls"):
            xls2xlsx = XLS2XLSX(self._document_path)
            self._document_path = f"{os.path.splitext(self._document_path)[0]}.xlsx"
            xls2xlsx.to_xlsx(self._document_path)

        self._workbook = load_workbook(self._document_path)
        self._worksheet = self._workbook["2022-2023"]

    def read_sheet(self) -> Dict:
        lesson_time = None
        day = None
        lessons = {}
        group_column_coordinates = self.get_groups_coordinates()
        for row in self._worksheet.iter_rows(min_row=7):
            for cell in row:
                if cell.value is None:
                    if cell.column in group_column_coordinates:
                        cell_value = "Нет пары"
                    else:
                        continue
                else:
                    cell_value = str(cell.value)
                if cell_value == '1' or cell_value == '3' or cell_value == '5' or cell_value == '7':
                    lesson_time = cell_value + f'-{int(cell_value) + 1} урок'
                    if lesson_time not in lessons[day]:
                        lessons[day][lesson_time] = []
                if cell_value.strip().lower() in ["понедельник", "вторник", "среда", "четверг", "пятница"]:
                    day = cell_value.strip()
                    if day not in lessons:
                        lessons[day] = {}
                else:
                    print(cell_value.strip(), day, lesson_time)
                    print(lessons)
                    if len(cell_value) > 1:
                        subject = re.sub(" +", " ", cell_value.strip())
                        lessons[day][lesson_time].append(subject)
        print(lessons)
        return lessons

    def get_shedule(self) -> None:
        groups = self.get_group_columns()

    def get_group_columns(self, group_row_index=6) -> List:
        groups = []
        for row in self._worksheet.iter_rows(min_row=group_row_index, max_row=group_row_index):
            for cell in row:
                cell_value = str(cell.value)
                if cell_value.count('-') >= 1:
                    if cell_value.count('-') > 1:
                        if cell_value.count(',') == 0:
                            cell_value = re.sub('  +', ',', cell_value)
                        cell_value = cell_value.replace(' ', '')
                        cell_value = cell_value.replace(',', ", ")
                    else:
                        cell_value = cell_value.replace(' ', '')
                    groups.append(cell_value)
        print(groups)
        return groups

    def get_groups_coordinates(self, group_row_index=6) -> List:
        group_coordinates = []
        for row in self._worksheet.iter_rows(min_row=group_row_index, max_row=group_row_index):
            for cell in row:
                if cell.value is None:
                    continue
                cell_value = cell.value
                if cell_value.count('-'):
                    group_coordinates.append(cell.column)
                    print(cell.column)
        return group_coordinates

    def get_shedule_for_each_group(self) -> Dict:
        groups = self.get_group_columns()
        raw_shedule = self.read_sheet()
        shedule = {}
        for group_index in range(len(groups)):
            lesson_index = group_index
            shedule[groups[group_index]] = {}
            for day in raw_shedule:
                shedule[groups[group_index]][day] = {}
                for time in raw_shedule[day]:
                    print(day, time, raw_shedule[day][time][lesson_index], groups[group_index])
                    shedule[groups[group_index]][day][time] = raw_shedule[day][time][lesson_index]
        print(shedule)
        return shedule
