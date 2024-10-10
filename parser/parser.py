import os.path

from xls2xlsx import XLS2XLSX
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict, Any
import re
# from utils.json_utils import json_converter, save_json
from parser.utils.json_utils import json_converter, save_json


class ScheduleParser:
    def __init__(self, document_path: str) -> None:
        self._document_path = document_path

    def _open_worksheet(self) -> None:
        if self._document_path.endswith(".xls"):
            xls2xlsx = XLS2XLSX(self._document_path)
            self._document_path = f"{os.path.splitext(self._document_path)[0]}.xlsx"
            xls2xlsx.to_xlsx(self._document_path)

        self._workbook = load_workbook(self._document_path)
        self._worksheet = self._workbook.worksheets[0]

    def get_group_row_index(self) -> int:
        index = 0
        for row in self._worksheet.iter_rows():
            index += 1
            for cell in row:
                if cell.value is not None:
                    # print(cell.value, row)
                    cell_value = str(cell.value)
                    if re.match(r'^[А-Яа-я]{1,3} - \d{2}', cell_value) is not None:
                        return index

    def get_groups(self, group_row_index: int) -> List:
        groups: List = []
        for row in self._worksheet.iter_rows(min_row=group_row_index, max_row=group_row_index):
            for cell in row:
                cell_value: str = str(cell.value)
                if cell_value.count('-') >= 1:
                    if cell_value.count('-') > 1:
                        if cell_value.count(',') == 0:
                            cell_value = re.sub('  +', ',', cell_value)
                        cell_value = cell_value.replace(' ', '')
                        cell_value = cell_value.replace(',', ", ")
                    else:
                        cell_value = cell_value.replace(' ', '')
                    groups.append(cell_value)
        print(groups)
        return groups

    def get_group_coordinates(self, group_row_index: int) -> List:
        study_group_coordinates: List = []
        for row in self._worksheet.iter_rows(min_row=group_row_index, max_row=group_row_index):
            for cell in row:
                if cell.value is None or cell.value.count('-') == 0:
                    continue
                else:
                    study_group_coordinates.append(cell.column)
        return study_group_coordinates

    def end_of_group_row(self, group_row_index) -> int:
        group_columns = self.get_group_coordinates(self.get_group_row_index())
        cell = self._worksheet.cell(group_row_index, group_columns[0])
        if self._worksheet.cell(group_row_index + 1, group_columns[0]).value is None:
            return group_row_index + 1
        else:
            return group_row_index

    def read_sheet(self) -> Dict:
        lesson_time: str | None = None
        day: str | None = None
        lessons: Dict = {}
        group_row_index: int = self.get_group_row_index()
        group_column_coordinates: List = self.get_group_coordinates(group_row_index)

        study_days: List = ["понедельник", "вторник", "среда", "четверг", "пятница", "суббота"]
        lesson_numbers: List = ['1', '3', '5', '7', '9', "11"]
        end_of_group_row = self.end_of_group_row(group_row_index)

        for row in self._worksheet.iter_rows(min_row=end_of_group_row + 1):
            for cell in row:
                cell_value: Any = cell.value
                if cell_value is None:
                    if cell.column in group_column_coordinates:
                        cell_value = "Нет пары"
                        # print(cell_value, cell.column, cell, row)
                    else:
                        continue
                else:
                    cell_value = str(cell.value)
                if cell_value in lesson_numbers:
                    lesson_time = cell_value + f'-{int(cell_value) + 1} урок'
                    if lesson_time not in lessons[day]:
                        lessons[day][lesson_time] = []
                elif any(day in cell_value.strip().lower() for day in study_days):
                    day = cell_value.strip()
                    if day not in lessons:
                        lessons[day] = {}
                else:
                    if len(cell_value) > 1:
                        subject: str = re.sub(" +", " ", cell_value.strip())
                        lessons[day][lesson_time].append(subject)
        print(lessons)
        return lessons

    def get_schedule_for_each_group(self) -> Dict:
        group_row_index = self.get_group_row_index()
        groups: List = self.get_groups(group_row_index)
        raw_schedule: Dict = self.read_sheet()
        schedule: Dict = {}
        for group_index in range(len(groups)):
            lesson_index: int = group_index
            schedule[groups[group_index]] = {}
            for day in raw_schedule:
                schedule[groups[group_index]][day] = {}
                for time in raw_schedule[day]:
                    schedule[groups[group_index]][day][time] = raw_schedule[day][time][lesson_index]
        print(schedule)
        return schedule

    @staticmethod
    def get_json_schedule(schedule: Dict, file_name) -> None:
        json_schedule = json_converter(schedule)
        print(json_schedule)
        save_json(json_schedule, file_name)

    @staticmethod
    def get_json_groups(groups: List, file_name) -> None:
        json_groups = json_converter(groups)
        print(json_groups)
        save_json(json_groups, file_name)
